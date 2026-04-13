"""
Potensi Claim Generator
Generates a Potensi Claim Excel file by merging:
  1. Breach Monitoring data (4 columns: Sisa Aging, Grouping Aging, Potensial Nominal Claim, Grouping Nominal)
  2. Apex Potensi Claim (M117) raw data (180 columns: AWB → TGL_TARIK_REPORT)
  3. Formula columns A–AK (CUST NAME → OPEN/CLOSE) taken from the CCC template and shifted +4 cols

The output layout is:
  Cols A-AK: Formula columns (37 cols)
  Cols AL-AO: Breach Monitoring data (4 cols)
  Cols AP onwards: Apex M117 data (180 cols)
"""
import os
import re
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formula.translate import Translator

# ------------- Config -----------------
UPLOADS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads")
APEX_M117_PATH = os.path.join(UPLOADS_DIR, "master_117", "master_data_117.xlsx")
BREACH_MONITORING_PATH = os.path.join(UPLOADS_DIR, "breach_monitoring", "breach_monitoring_data.xlsx")
CCC_TEMPLATE_PATH = os.path.join(UPLOADS_DIR, "db_ccc", "db_ccc_data.xlsx")
OUTPUT_PATH = os.path.join(UPLOADS_DIR, "Template_Potensi_Claim.xlsx")

# Breach Monitoring columns we need
BREACH_COLS = ["Sisa Aging", "Grouping Aging", "Potensial Nominal Claim", "Grouping Nominal"]
BREACH_KEY = "Tracking Number"  # This maps to AWB in M117

# In CCC template, Apex data starts at column AL (col 38)
CCC_APEX_START_COL = 38  # AL

# In our output, Breach data occupies AL-AO (cols 38-41), and Apex starts at AP (col 42)
OUTPUT_APEX_START_COL = 42  # AP

# The shift: all formula references to Apex columns need to move right by 4
COL_SHIFT = OUTPUT_APEX_START_COL - CCC_APEX_START_COL  # 4


def _shift_col_letter(col_letter, shift):
    """Shift a column letter by a number of positions. E.g., 'AL' + 4 = 'AP'."""
    col_idx = column_index_from_string(col_letter)
    new_idx = col_idx + shift
    return get_column_letter(new_idx)


def _shift_formula_refs(formula, shift, min_col=38):
    """
    Shift cell references in a formula that point to columns >= min_col by `shift` positions.
    E.g., AL2 -> AP2, AM2 -> AQ2, etc. References to columns A-AK (1-37) are NOT shifted.
    """
    # Match cell references like AL2, AM$3, $CE2, etc.
    # We need to handle both relative and absolute column references
    pattern = r'(\$?)([A-Z]{1,3})(\$?)(\d+)'
    
    def replace_ref(match):
        abs_col = match.group(1)  # $ or empty
        col_letters = match.group(2)
        abs_row = match.group(3)  # $ or empty
        row_num = match.group(4)
        
        # Don't shift if column is absolute ($) - these are usually sheet references
        if abs_col == '$':
            return match.group(0)
        
        try:
            col_idx = column_index_from_string(col_letters)
        except ValueError:
            return match.group(0)
        
        # Only shift columns that are in the Apex data range (>= min_col)
        if col_idx >= min_col:
            new_col_idx = col_idx + shift
            new_col_letters = get_column_letter(new_col_idx)
            return f"{abs_col}{new_col_letters}{abs_row}{row_num}"
        
        return match.group(0)
    
    return re.sub(pattern, replace_ref, formula)


def generate_potensi_claim(output_path=None):
    """
    Generate the Potensi Claim Excel file.
    Returns the output path on success.
    """
    if output_path is None:
        output_path = OUTPUT_PATH
    
    # 1. Check if required files exist
    if not os.path.exists(APEX_M117_PATH):
        raise FileNotFoundError(f"Apex Potensi Claim (M117) file not found: {APEX_M117_PATH}")
    if not os.path.exists(BREACH_MONITORING_PATH):
        raise FileNotFoundError(f"Breach Monitoring file not found: {BREACH_MONITORING_PATH}")
    if not os.path.exists(CCC_TEMPLATE_PATH):
        raise FileNotFoundError(f"CCC Template file not found: {CCC_TEMPLATE_PATH}")
    
    print("[PotensiClaim] Loading data files...")
    
    # 2. Load Apex M117 data
    df_apex = pd.read_excel(APEX_M117_PATH)
    df_apex = df_apex.loc[:, ~df_apex.columns.str.contains('^Unnamed')]
    print(f"[PotensiClaim] Apex M117: {len(df_apex)} rows, {len(df_apex.columns)} cols")
    
    if len(df_apex) == 0:
        raise ValueError("Apex M117 data is empty")
    
    # 3. Load Breach Monitoring data (it's a CSV disguised as .xlsx, latin-1 encoded)
    try:
        df_breach = pd.read_csv(BREACH_MONITORING_PATH, encoding='latin-1')
    except Exception:
        # Fallback: try as a real xlsx
        try:
            df_breach = pd.read_excel(BREACH_MONITORING_PATH)
        except Exception:
            df_breach = pd.read_excel(BREACH_MONITORING_PATH, engine='xlrd')
    
    df_breach = df_breach.loc[:, ~df_breach.columns.str.contains('^Unnamed')]
    print(f"[PotensiClaim] Breach Monitoring: {len(df_breach)} rows, {len(df_breach.columns)} cols")
    
    # Build a lookup from Breach data: AWB -> {Sisa Aging, Grouping Aging, ...}
    breach_lookup = {}
    if BREACH_KEY in df_breach.columns:
        for _, row in df_breach.iterrows():
            awb = str(row[BREACH_KEY]).strip()
            breach_lookup[awb] = {
                col: row[col] if col in df_breach.columns else None
                for col in BREACH_COLS
            }
    print(f"[PotensiClaim] Breach lookup: {len(breach_lookup)} entries")
    
    # 4. Load CCC template to extract formulas from DATA sheet row 2
    wb_ccc = openpyxl.load_workbook(CCC_TEMPLATE_PATH)
    ws_ccc = wb_ccc['DATA']
    
    # Extract formulas from columns A-AK (1-37) in row 2
    ccc_formulas = {}
    for c in range(1, 38):  # cols 1-37 = A-AK
        cell_val = ws_ccc.cell(row=2, column=c).value
        if cell_val and str(cell_val).startswith('='):
            # Shift Apex column references by +4
            shifted = _shift_formula_refs(str(cell_val), COL_SHIFT, min_col=CCC_APEX_START_COL)
            ccc_formulas[c] = shifted
    
    print(f"[PotensiClaim] Extracted {len(ccc_formulas)} formulas from CCC template")
    
    # Also extract headers from CCC template for formula columns
    ccc_headers = {}
    for c in range(1, 38):
        val = ws_ccc.cell(row=1, column=c).value
        if val:
            ccc_headers[c] = val
    
    wb_ccc.close()
    
    # 5. Create the output workbook
    # We'll use the CCC template as a base to preserve reference sheets
    wb_out = openpyxl.load_workbook(CCC_TEMPLATE_PATH)
    ws_out = wb_out['DATA']
    
    # Clear existing data rows (keep row 1 header)
    if ws_out.max_row > 1:
        ws_out.delete_rows(2, ws_out.max_row)
    
    # 6. Write headers for row 1
    # Cols A-AK: formula column headers (already there from template, but let's ensure)
    for c, header in ccc_headers.items():
        ws_out.cell(row=1, column=c).value = header
    
    # Cols AL-AO: Breach data headers
    for i, breach_col in enumerate(BREACH_COLS):
        ws_out.cell(row=1, column=38 + i).value = breach_col  # AL=38, AM=39, AN=40, AO=41
    
    # Cols AP onwards: Apex M117 headers
    apex_columns = list(df_apex.columns)
    for i, col_name in enumerate(apex_columns):
        ws_out.cell(row=1, column=OUTPUT_APEX_START_COL + i).value = col_name  # AP=42, AQ=43, ...
    
    # 7. Write data rows
    apex_records = df_apex.to_dict('records')
    num_rows = len(apex_records)
    print(f"[PotensiClaim] Writing {num_rows} rows...")
    
    for row_idx, record in enumerate(apex_records, start=2):
        awb = str(record.get('AWB', '')).strip()
        
        # Write Breach data (cols AL-AO = 38-41)
        breach_data = breach_lookup.get(awb, {})
        for i, breach_col in enumerate(BREACH_COLS):
            val = breach_data.get(breach_col)
            if pd.notna(val) if not isinstance(val, type(None)) else False:
                ws_out.cell(row=row_idx, column=38 + i).value = val
        
        # Write Apex data (cols AP onwards = 42+)
        for i, col_name in enumerate(apex_columns):
            val = record.get(col_name)
            if pd.notna(val) if not isinstance(val, type(None)) else False:
                ws_out.cell(row=row_idx, column=OUTPUT_APEX_START_COL + i).value = val
        
        # Write formulas (cols A-AK = 1-37)
        for calc_col, formula in ccc_formulas.items():
            if row_idx == 2:
                # First data row: use the shifted formula directly
                ws_out.cell(row=row_idx, column=calc_col).value = formula
            else:
                # Translate formula from row 2 to current row
                origin_cell = f"{get_column_letter(calc_col)}2"
                dest_cell = f"{get_column_letter(calc_col)}{row_idx}"
                try:
                    translator = Translator(formula, origin=origin_cell)
                    translated = translator.translate_formula(dest=dest_cell)
                    ws_out.cell(row=row_idx, column=calc_col).value = translated
                except Exception:
                    # If translation fails, just copy the formula with row adjustment
                    ws_out.cell(row=row_idx, column=calc_col).value = formula.replace('2,', f'{row_idx},').replace('2)', f'{row_idx})')
    
    # 8. Save
    wb_out.save(output_path)
    print(f"[PotensiClaim] Saved to {output_path}")
    return output_path


if __name__ == "__main__":
    import traceback
    try:
        result = generate_potensi_claim()
        print(f"Success! Output: {result}")
    except Exception as e:
        print(f"Error: {e}")
        traceback.print_exc()
