import os
import shutil
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def update_sheet_with_df(wb, sheet_name, df):
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.delete_rows(1, sheet.max_row)
        for r in dataframe_to_rows(df, index=False, header=True):
            sheet.append(r)

def load_reference(ref_path):
    if os.path.exists(ref_path):
        return pd.read_excel(ref_path)
    return pd.DataFrame()

def generate_template_ccc(master_data, output_path, template_path, refs_dir):
    """
    Generates output Excel by:
    1. Copying the template byte-for-byte (preserves ALL formulas, styles, named ranges)
    2. Updating reference sheets
    3. Writing master Apex data ONLY into Apex columns (col 38+)
    
    Formula strings are NEVER written by openpyxl to avoid XML corruption
    caused by unescaped characters (<, >, &) in formula text literals.
    """
    # Step 1: Copy template byte-for-byte to output path
    shutil.copy2(template_path, output_path)

    # Step 2: Open the copy and update reference sheets
    wb = openpyxl.load_workbook(output_path, keep_links=False)

    ref_map = {
        "Database SLA Shopee": os.path.join(refs_dir, "sla_shopee", "ref_sla_shopee.xlsx"),
        "Database SLA Lazada": os.path.join(refs_dir, "sla_lazada", "ref_sla_lazada.xlsx"),
        "Database 1": os.path.join(refs_dir, "kategori_inbound", "ref_kategori_inbound.xlsx"),
        "Database 2": os.path.join(refs_dir, "db_2", "ref_db_2.xlsx"),
        "SERVICE": os.path.join(refs_dir, "kategori_service", "ref_kategori_service.xlsx"),
        "Account update  06-01-2026": os.path.join(refs_dir, "account_update", "ref_account_update.xlsx"),
        "Sheet1": os.path.join(refs_dir, "tabel_coding", "ref_tabel_coding.xlsx"),
    }

    for sheet_name, ref_path in ref_map.items():
        df_ref = load_reference(ref_path)
        if not df_ref.empty:
            df_ref = df_ref.loc[:, ~df_ref.columns.str.contains('^Unnamed')]
            update_sheet_with_df(wb, sheet_name, df_ref)

    # Step 3: Load master data
    if isinstance(master_data, pd.DataFrame):
        df_master = master_data.copy()
    else:
        df_master = pd.read_excel(master_data)

    df_master = df_master.loc[:, ~df_master.columns.str.contains('^Unnamed')]
    master_columns = list(df_master.columns)
    num_rows = len(df_master)

    if num_rows == 0:
        wb.save(output_path)
        return output_path

    # Step 4: Write master Apex data into DATA sheet — Apex cols only (38+)
    # NEVER write formula strings — openpyxl corrupts xlsx when formulas contain <, >, & characters
    sheet = wb["DATA"]
    APEX_START_COL = 38

    # Write Apex column headers into row 1
    for i, col_name in enumerate(master_columns):
        sheet.cell(row=1, column=APEX_START_COL + i).value = col_name

    # Write data rows — only Apex data columns, formula cols (1-37) untouched
    for row_idx in range(num_rows):
        excel_row = row_idx + 2
        for i, col_name in enumerate(master_columns):
            val = df_master.iat[row_idx, i]
            # Convert NaN/NaT to None (blank cell) — never to empty string
            try:
                if pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            sheet.cell(row=excel_row, column=APEX_START_COL + i).value = val

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    import traceback
    try:
        master_data_path = r"e:\BPS JNE DASHBOARD\backend\uploads\master\master_data.xlsx"
        template_path = r"e:\BPS JNE DASHBOARD\backend\uploads\db_ccc\db_ccc_data.xlsx"
        output_path = r"C:\tmp\test_template_ccc.xlsx"
        refs_dir = r"e:\BPS JNE DASHBOARD\backend\uploads\referensi"
        generate_template_ccc(master_data_path, output_path, template_path, refs_dir)
        print("Success! Saved to", output_path)
    except Exception as e:
        print("Error:")
        traceback.print_exc()
